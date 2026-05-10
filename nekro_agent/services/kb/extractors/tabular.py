from __future__ import annotations

import csv
import json
from pathlib import Path

import json5
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .base import ExtractedKBText, clean_text


def extract_json_like(file_path: Path) -> ExtractedKBText:
    raw = file_path.read_text(encoding="utf-8", errors="replace")
    try:
        parsed = json5.loads(raw)
        pretty = json.dumps(parsed, ensure_ascii=False, indent=2)
        return ExtractedKBText(text=clean_text(pretty))
    except Exception:
        return ExtractedKBText(text=clean_text(raw))


def extract_yaml_like(file_path: Path) -> ExtractedKBText:
    return ExtractedKBText(text=clean_text(file_path.read_text(encoding="utf-8", errors="replace")))


def extract_csv(file_path: Path) -> ExtractedKBText:
    rows_text: list[str] = []
    with file_path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        return ExtractedKBText(text="")
    header = rows[0]
    rows_text.append("# CSV Table")
    rows_text.append("")
    rows_text.append("## Header")
    rows_text.append(", ".join(header))
    for index, row in enumerate(rows[1:], start=1):
        rows_text.append("")
        rows_text.append(f"## Row {index}")
        for key, value in zip(header, row, strict=False):
            rows_text.append(f"- {key}: {value}")
    return ExtractedKBText(text=clean_text("\n".join(rows_text)))


def extract_xlsx(file_path: Path) -> ExtractedKBText:
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        lines: list[str] = []

        for sheet in workbook.worksheets:
            raw_rows = [
                ["" if value is None else str(value).strip() for value in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            rows = [row for row in raw_rows if any(cell for cell in row)]
            if not rows:
                continue

            header = rows[0]
            header_labels = [
                cell if cell else f"Column {get_column_letter(index)}"
                for index, cell in enumerate(header, start=1)
            ]

            lines.append(f"# Sheet: {sheet.title}")
            lines.append("")
            lines.append("## Header")
            lines.append(", ".join(header_labels))

            for row_index, row in enumerate(rows[1:], start=1):
                if not any(cell for cell in row):
                    continue
                lines.append("")
                lines.append(f"## Row {row_index}")
                for column_index, value in enumerate(row, start=1):
                    if not value:
                        continue
                    key = (
                        header_labels[column_index - 1]
                        if column_index - 1 < len(header_labels)
                        else f"Column {get_column_letter(column_index)}"
                    )
                    lines.append(f"- {key}: {value}")

            lines.append("")

        return ExtractedKBText(text=clean_text("\n".join(lines)))
    finally:
        workbook.close()
